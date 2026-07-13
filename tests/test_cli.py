import contextlib
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from importlib.metadata import PackageNotFoundError, version
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from excelflow import create_template
from excelflow.cli import main, package_version

ROOT = Path(__file__).resolve().parents[1]
PLAN = ROOT / "examples/tutorial/01_single_sheet.xlsx"
SOURCE = ROOT / "examples/tutorial/source.xlsx"
EXPECTED_RECORDS = [
    {"order_id": 1001, "status": "paid", "amount": 727.0},
    {"order_id": 1002, "status": "pending", "amount": 499.0},
    {"order_id": 1003, "status": "paid", "amount": 80.0},
]


class CliTest(unittest.TestCase):
    def run_cli(self, *args, cwd=None):
        stdout, stderr = io.StringIO(), io.StringIO()
        original_cwd = Path.cwd()
        try:
            if cwd is not None:
                os.chdir(cwd)
            with (
                patch.object(sys, "argv", ["excelflow", *args]),
                contextlib.redirect_stdout(stdout),
                contextlib.redirect_stderr(stderr),
            ):
                try:
                    returncode = main()
                except SystemExit as exc:
                    returncode = int(exc.code or 0)
        finally:
            os.chdir(original_cwd)
        return SimpleNamespace(
            returncode=returncode, stdout=stdout.getvalue(), stderr=stderr.getvalue()
        )

    def assert_parse_error(self, result):
        self.assertEqual(result.returncode, 2)
        self.assertIn("usage:", result.stderr)
        self.assertEqual(result.stdout, "")

    def test_global_and_subcommand_help(self):
        top = self.run_cli("--help")
        self.assertEqual(top.returncode, 0)
        for command in ("template", "validate", "preview", "run"):
            self.assertIn(command, top.stdout)
            result = self.run_cli(command, "--help")
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertIn(f"excelflow {command}", result.stdout)
        self.assertIn("-o OUTPUT, --output OUTPUT", self.run_cli("template", "--help").stdout)
        self.assertIn("-p PLAN, --plan PLAN", self.run_cli("validate", "--help").stdout)
        preview_help = self.run_cli("preview", "--help").stdout
        self.assertIn("-p PLAN, --plan PLAN", preview_help)
        self.assertIn("-t TASK, --task TASK", preview_help)
        run_help = self.run_cli("run", "--help").stdout
        for option in ("--plan", "--task", "--source", "--format", "--output"):
            self.assertIn(option, run_help)

    def test_version_long_short_console_and_module_entrypoints(self):
        expected = f"excelflow {version('excelflow')}"
        for option in ("--version", "-V"):
            result = self.run_cli(option)
            self.assertEqual(result.returncode, 0)
            self.assertEqual(result.stdout.strip(), expected)
            self.assertEqual(result.stderr, "")
        executable = shutil.which("excelflow")
        self.assertIsNotNone(executable, "console entrypoint is not installed on PATH")
        console = subprocess.run(
            [executable, "--version"], capture_output=True, text=True, check=False
        )
        self.assertEqual(console.returncode, 0, console.stderr)
        self.assertEqual(console.stdout.strip(), expected)
        for option in ("--help", "--version"):
            module = subprocess.run(
                [sys.executable, "-m", "excelflow", option],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(module.returncode, 0, module.stderr)
            self.assertIn("excelflow", module.stdout)

    def test_version_fallback_without_distribution_metadata(self):
        with patch("excelflow.cli.version", side_effect=PackageNotFoundError):
            self.assertEqual(package_version(), "0.0.0+unknown")

    def test_template_default_long_and_short_output(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            default = self.run_cli("template", cwd=root)
            self.assertEqual(default.returncode, 0, default.stderr)
            self.assertEqual(default.stderr, "")
            self.assertTrue((root / "extraction_plan.xlsx").exists())
            for option, name in (("--output", "long.xlsx"), ("-o", "short.xlsx")):
                result = self.run_cli("template", option, str(root / name))
                self.assertEqual(result.returncode, 0, result.stderr)
                self.assertEqual(result.stderr, "")
                self.assertTrue((root / name).exists())

    def test_validate_long_and_short_plan(self):
        for option in ("--plan", "-p"):
            result = self.run_cli("validate", option, str(PLAN))
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertEqual(result.stdout.strip(), "校验通过")
            self.assertEqual(result.stderr, "")

    def test_preview_long_and_short_options(self):
        commands = (
            ("--plan", str(PLAN), "--task", "lesson_01"),
            ("-p", str(PLAN), "-t", "lesson_01"),
        )
        for args in commands:
            result = self.run_cli("preview", *args)
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertIn("Pandas 执行计划: lesson_01", result.stdout)
            self.assertIn("读取: 订单 -> o", result.stdout)
            self.assertEqual(result.stderr, "")

    def test_run_long_options_to_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "result.csv"
            result = self.run_cli(
                "run",
                "--plan",
                str(PLAN),
                "--task",
                "lesson_01",
                "--source",
                str(SOURCE),
                "--format",
                "csv",
                "--output",
                str(output),
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertEqual(result.stderr, "")
            self.assertEqual(pd.read_csv(output).to_dict("records"), EXPECTED_RECORDS)
            self.assertIn("抽取完成: 3 行", result.stdout)

    def test_run_short_options_for_every_output_format(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            for output_format in ("csv", "jsonl", "xlsx"):
                output = root / f"result.{output_format}"
                result = self.run_cli(
                    "run",
                    "-p",
                    str(PLAN),
                    "-t",
                    "lesson_01",
                    "-s",
                    str(SOURCE),
                    "-f",
                    output_format,
                    "-o",
                    str(output),
                )
                self.assertEqual(result.returncode, 0, result.stderr)
                self.assertEqual(result.stderr, "")
                self.assertTrue(output.exists())
                if output_format == "csv":
                    records = pd.read_csv(output).to_dict("records")
                elif output_format == "jsonl":
                    records = [json.loads(x) for x in output.read_text().splitlines()]
                else:
                    records = pd.read_excel(output).to_dict("records")
                self.assertEqual(records, EXPECTED_RECORDS)

    def test_run_defaults_format_and_output_when_omitted(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            # 省略 --format 与 --output：默认 xlsx，文件名 <task>.<格式>
            result = self.run_cli(
                "run",
                "-p",
                str(PLAN),
                "-t",
                "lesson_01",
                "-s",
                str(SOURCE),
                cwd=root,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertEqual(result.stderr, "")
            default_output = root / "lesson_01.xlsx"
            self.assertTrue(default_output.exists())
            self.assertEqual(pd.read_excel(default_output).to_dict("records"), EXPECTED_RECORDS)
            self.assertIn("抽取完成: 3 行", result.stdout)
            # 仅省略 --output：文件名后缀跟随 --format
            result_csv = self.run_cli(
                "run",
                "-p",
                str(PLAN),
                "-t",
                "lesson_01",
                "-s",
                str(SOURCE),
                "-f",
                "csv",
                cwd=root,
            )
            self.assertEqual(result_csv.returncode, 0, result_csv.stderr)
            self.assertTrue((root / "lesson_01.csv").exists())

    @staticmethod
    def _make_multi_task_plan(root: Path) -> Path:
        source = root / "source.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "订单"
        ws.append(["id", "amount"])
        ws.append([1, 10])
        ws.append([2, 20])
        wb.save(source)
        plan_path = root / "plan.xlsx"
        create_template(plan_path)
        plan_wb = load_workbook(plan_path)
        plan = plan_wb["抽取计划"]
        plan.delete_rows(2, plan.max_row)
        plan.append(["t1", "是", ""])
        plan.append(["t2", "是", ""])
        plan.append(["t3", "否", "未启用，应被跳过"])
        objects = plan_wb["数据对象"]
        objects.delete_rows(2, objects.max_row)
        # 三个任务都配主表，保持结构合法（validator 不区分启用与否）
        for task_id in ("t1", "t2", "t3"):
            objects.append([task_id, "订单", "o", 1, "是", ""])
        plan_wb["关联关系"].delete_rows(2, plan_wb["关联关系"].max_row)  # 清默认 demo 行
        fields = plan_wb["字段映射"]
        fields.delete_rows(2, fields.max_row)
        fields.append(["t1", "o.id", "order_id", "integer", "", 1, ""])
        fields.append(["t2", "o.id", "order_id", "integer", "", 1, ""])
        fields.append(["t2", "o.amount", "amount", "decimal", "", 2, ""])
        fields.append(["t3", "o.id", "order_id", "integer", "", 1, ""])
        plan_wb["过滤条件"].delete_rows(2, plan_wb["过滤条件"].max_row)  # 清默认 demo 行
        plan_wb.save(plan_path)
        return plan_path

    def test_run_all_executes_every_enabled_task(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            plan_path = self._make_multi_task_plan(root)
            out_dir = root / "out"
            result = self.run_cli(
                "run",
                "-p",
                str(plan_path),
                "-s",
                str(root / "source.xlsx"),
                "--format",
                "csv",
                "--output",
                str(out_dir),
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue((out_dir / "t1.csv").exists())
            self.assertTrue((out_dir / "t2.csv").exists())
            self.assertFalse((out_dir / "t3.csv").exists())
            self.assertIn("[t1]", result.stdout)
            self.assertIn("[t2]", result.stdout)
            self.assertNotIn("[t3]", result.stdout)

    def test_run_all_rejects_output_pointing_at_a_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            plan_path = self._make_multi_task_plan(root)
            bad_output = root / "not_a_dir.csv"
            bad_output.write_text("x")
            result = self.run_cli(
                "run",
                "-p",
                str(plan_path),
                "-s",
                str(root / "source.xlsx"),
                "-o",
                str(bad_output),
            )
            self.assertEqual(result.returncode, 1)
            self.assertIn("必须是目录", result.stderr)

    def test_run_all_wraps_failure_with_task_id_and_keeps_prior_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "订单"
            ws.append(["id"])
            ws.append([1])
            wb.save(source)
            plan_path = root / "plan.xlsx"
            create_template(plan_path)
            plan_wb = load_workbook(plan_path)
            plan_wb["抽取计划"].delete_rows(2, plan_wb["抽取计划"].max_row)
            plan_wb["抽取计划"].append(["t1", "是", ""])
            plan_wb["抽取计划"].append(["t2", "是", ""])
            objects = plan_wb["数据对象"]
            objects.delete_rows(2, objects.max_row)
            objects.append(["t1", "订单", "o", 1, "是", ""])
            objects.append(["t2", "订单X", "o", 1, "是", ""])  # 订单X 不在 source
            plan_wb["关联关系"].delete_rows(2, plan_wb["关联关系"].max_row)
            fields = plan_wb["字段映射"]
            fields.delete_rows(2, fields.max_row)
            fields.append(["t1", "o.id", "order_id", "integer", "", 1, ""])
            fields.append(["t2", "o.id", "order_id", "integer", "", 1, ""])
            plan_wb["过滤条件"].delete_rows(2, plan_wb["过滤条件"].max_row)
            plan_wb.save(plan_path)
            out_dir = root / "out"
            result = self.run_cli(
                "run",
                "-p",
                str(plan_path),
                "-s",
                str(source),
                "--format",
                "csv",
                "--output",
                str(out_dir),
            )
            # 多任务失败时异常带 task_id 前缀；fail-fast，但 t1 已写出的文件保留
            self.assertEqual(result.returncode, 1)
            self.assertIn("任务 t2 失败", result.stderr)
            self.assertTrue((out_dir / "t1.csv").exists())
            self.assertFalse((out_dir / "t2.csv").exists())

    def test_old_positional_syntax_is_rejected_for_every_command(self):
        commands = (
            ("template", "plan.xlsx"),
            ("validate", "plan.xlsx"),
            ("preview", "plan.xlsx", "task"),
            ("run", "plan.xlsx", "task", "source.xlsx", "csv", "output.csv"),
        )
        for args in commands:
            with self.subTest(command=args[0]):
                self.assert_parse_error(self.run_cli(*args))

    def test_every_required_option_is_enforced(self):
        commands = [
            ("validate",),
            ("preview", "--task", "lesson_01"),
            ("preview", "--plan", str(PLAN)),
        ]
        run_pairs = [
            ("--plan", str(PLAN)),
            ("--source", str(SOURCE)),
        ]
        for missing_index in range(len(run_pairs)):
            args = [
                value
                for index, pair in enumerate(run_pairs)
                if index != missing_index
                for value in pair
            ]
            commands.append(("run", *args))
        for args in commands:
            with self.subTest(args=args):
                self.assert_parse_error(self.run_cli(*args))

    def test_invalid_unknown_and_abbreviated_options_are_rejected(self):
        commands = (
            ("unknown",),
            ("validate", "--pla", str(PLAN)),
            (
                "run",
                "-p",
                str(PLAN),
                "-t",
                "lesson_01",
                "-s",
                str(SOURCE),
                "-f",
                "xml",
                "-o",
                "out.xml",
            ),
        )
        for args in commands:
            with self.subTest(args=args):
                self.assert_parse_error(self.run_cli(*args))

    def test_runtime_failures_use_exit_one_and_stderr(self):
        missing = self.run_cli("validate", "--plan", "/missing/plan.xlsx")
        self.assertEqual(missing.returncode, 1)
        self.assertEqual(missing.stdout, "")
        self.assertIn("错误:", missing.stderr)
        unknown_task = self.run_cli("preview", "--plan", str(PLAN), "--task", "missing")
        self.assertEqual(unknown_task.returncode, 1)
        self.assertEqual(unknown_task.stdout, "")
        self.assertIn("失败: 任务不存在", unknown_task.stderr)
        disabled = self.run_cli(
            "run",
            "-p",
            str(ROOT / "extraction_plan.xlsx"),
            "-t",
            "demo_orders",
            "-s",
            "/missing/source.xlsx",
            "-f",
            "csv",
            "-o",
            "/tmp/disabled.csv",
        )
        self.assertEqual(disabled.returncode, 1)
        self.assertEqual(disabled.stdout, "")
        self.assertIn("未启用", disabled.stderr)
        missing_source = self.run_cli(
            "run",
            "-p",
            str(PLAN),
            "-t",
            "lesson_01",
            "-s",
            "/missing/source.xlsx",
            "-f",
            "csv",
            "-o",
            "/tmp/missing.csv",
        )
        self.assertEqual(missing_source.returncode, 1)
        self.assertEqual(missing_source.stdout, "")
        self.assertIn("失败:", missing_source.stderr)


if __name__ == "__main__":
    unittest.main()
