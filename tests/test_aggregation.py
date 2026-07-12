import tempfile
import unittest
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from excelflow.engine import PandasExtractionEngine
from excelflow.repository import ExcelSpecRepository
from excelflow.schema import AGGREGATION_HEADERS, GROUP_HEADERS, ExtractionSpec
from excelflow.service import ExtractionService
from excelflow.template import create_template
from excelflow.validator import SpecValidator


class AggregationTest(unittest.TestCase):
    def setUp(self):
        self.engine = PandasExtractionEngine()
        self.frame = pd.DataFrame({
            "o.department": ["A", "A", "B", None],
            "o.id": [1, 1, 2, 3], "o.amount": [10.0, None, 5.0, 7.0],
            "o.label": ["x", None, "y", "z"],
        })
        self.groups = [{"任务ID": "summary", "源字段": "o.department", "目标字段": "department", "目标类型": "string", "分组顺序": 1}]

    def rule(self, source, function, target, target_type, order, separator=None):
        return {"任务ID": "summary", "源字段": source, "聚合函数": function, "目标字段": target,
                "目标类型": target_type, "分隔符": separator, "聚合顺序": order}

    def test_all_grouped_aggregations_and_null_group(self):
        rules = [
            self.rule("", "count_all", "rows", "integer", 1),
            self.rule("o.amount", "count", "amount_count", "integer", 2),
            self.rule("o.id", "count_distinct", "orders", "integer", 3),
            self.rule("o.amount", "sum", "total", "decimal", 4),
            self.rule("o.amount", "avg", "average", "decimal", 5),
            self.rule("o.amount", "min", "minimum", "decimal", 6),
            self.rule("o.amount", "max", "maximum", "decimal", 7),
            self.rule("o.label", "first", "first_label", "string", 8),
            self.rule("o.label", "last", "last_label", "string", 9),
            self.rule("o.label", "concat_agg", "labels", "string", 10, "|")]
        result = self.engine._aggregate(self.frame, ExtractionSpec(groups=self.groups, aggregations=rules), "summary")
        self.assertEqual(result.columns.tolist(), ["department", "rows", "amount_count", "orders", "total", "average", "minimum", "maximum", "first_label", "last_label", "labels"])
        self.assertEqual(result["department"].tolist()[:3], ["A", "B", pd.NA])
        self.assertEqual(result["rows"].tolist(), [2, 1, 1])
        self.assertEqual(result["amount_count"].tolist(), [1, 1, 1])
        self.assertEqual(result["orders"].tolist(), [1, 1, 1])
        self.assertEqual(result["total"].tolist(), [10.0, 5.0, 7.0])
        self.assertEqual(result["labels"].tolist(), ["x", "y", "z"])

    def test_grouped_value_aggregations_are_asserted(self):
        # The grouped path uses series_group.mean()/groupby.first()/last(), distinct from the
        # whole-table APIs; these values must be locked so a wrong function or dropped NA-skip
        # cannot pass silently. Group A (amount [10.0, None], label ["x", None]) skips NA.
        rules = [
            self.rule("o.amount", "avg", "average", "decimal", 1),
            self.rule("o.amount", "min", "minimum", "decimal", 2),
            self.rule("o.amount", "max", "maximum", "decimal", 3),
            self.rule("o.label", "first", "first_label", "string", 4),
            self.rule("o.label", "last", "last_label", "string", 5)]
        result = self.engine._aggregate(self.frame, ExtractionSpec(groups=self.groups, aggregations=rules), "summary")
        self.assertEqual(result["average"].tolist(), [10.0, 5.0, 7.0])
        self.assertEqual(result["minimum"].tolist(), [10.0, 5.0, 7.0])
        self.assertEqual(result["maximum"].tolist(), [10.0, 5.0, 7.0])
        # last_label == "x" for group A (not None) locks the NA-skipping semantics of groupby.last().
        self.assertEqual(result["first_label"].tolist(), ["x", "y", "z"])
        self.assertEqual(result["last_label"].tolist(), ["x", "y", "z"])

    def test_multi_key_grouping_combines_columns(self):
        frame = pd.DataFrame({
            "o.region": ["东", "东", "西", "西", "东"],
            "o.dept": ["A", "B", "A", "A", "A"],
            "o.amount": [10.0, 20.0, 30.0, 40.0, 50.0],
        })
        groups = [
            {"任务ID": "summary", "源字段": "o.region", "目标字段": "region", "目标类型": "string", "分组顺序": 1},
            {"任务ID": "summary", "源字段": "o.dept", "目标字段": "dept", "目标类型": "string", "分组顺序": 2}]
        rules = [self.rule("", "count_all", "rows", "integer", 1), self.rule("o.amount", "sum", "total", "decimal", 2)]
        result = self.engine._aggregate(frame, ExtractionSpec(groups=groups, aggregations=rules), "summary")
        # Composite (region, dept) groups keep first-appearance order: (东,A), (东,B), (西,A).
        self.assertEqual(result.columns.tolist(), ["region", "dept", "rows", "total"])
        self.assertEqual(len(result), 3)
        self.assertEqual(result["region"].tolist(), ["东", "东", "西"])
        self.assertEqual(result["dept"].tolist(), ["A", "B", "A"])
        self.assertEqual(result["rows"].tolist(), [2, 1, 2])
        self.assertEqual(result["total"].tolist(), [60.0, 20.0, 70.0])

    def test_whole_table_first_last_skips_na_and_returns_na_when_empty(self):
        # Whole-table path (engine.py L135) is the structural twin of the grouped first/last just
        # locked. Existing data puts None in the middle, so dropping .dropna() would still pass;
        # None at the head/tail is the shape that truly distinguishes skip-na behavior.
        framed = pd.DataFrame({"o.tag": [None, "y", None]})
        result = self.engine._aggregate(framed, ExtractionSpec(aggregations=[
            self.rule("o.tag", "first", "first", "string", 1),
            self.rule("o.tag", "last", "last", "string", 2)]), "summary")
        self.assertEqual(result["first"].tolist(), ["y"])
        self.assertEqual(result["last"].tolist(), ["y"])
        # An all-null column must hit the else pd.NA guard, not raise IndexError on iloc[0].
        empty = pd.DataFrame({"o.tag": [None, None]})
        empty_result = self.engine._aggregate(empty, ExtractionSpec(aggregations=[
            self.rule("o.tag", "first", "first", "string", 1),
            self.rule("o.tag", "last", "last", "string", 2)]), "summary")
        self.assertTrue(pd.isna(empty_result["first"].iloc[0]))
        self.assertTrue(pd.isna(empty_result["last"].iloc[0]))

    def test_whole_table_aggregation_and_all_null_sum(self):
        frame = pd.DataFrame({"o.amount": [None, None], "o.id": [1, 2]})
        rules = [self.rule("", "count_all", "rows", "integer", 1), self.rule("o.amount", "sum", "total", "decimal", 2)]
        result = self.engine._aggregate(frame, ExtractionSpec(aggregations=rules), "summary")
        self.assertEqual(result["rows"].iloc[0], 2)
        self.assertTrue(pd.isna(result["total"].iloc[0]))

    def test_every_whole_table_aggregation(self):
        rules = [
            self.rule("", "count_all", "rows", "integer", 1),
            self.rule("o.amount", "count", "count", "integer", 2),
            self.rule("o.id", "count_distinct", "distinct", "integer", 3),
            self.rule("o.amount", "avg", "average", "decimal", 4),
            self.rule("o.amount", "min", "minimum", "decimal", 5),
            self.rule("o.amount", "max", "maximum", "decimal", 6),
            self.rule("o.label", "first", "first", "string", 7),
            self.rule("o.label", "last", "last", "string", 8),
            self.rule("o.label", "concat_agg", "labels", "string", 9, "|")]
        result = self.engine._aggregate(self.frame, ExtractionSpec(aggregations=rules), "summary").iloc[0]
        self.assertEqual(result.to_dict(), {"rows": 4, "count": 3, "distinct": 3, "average": 22 / 3,
            "minimum": 5.0, "maximum": 10.0, "first": "x", "last": "z", "labels": "x|y|z"})

    def test_sum_rejects_non_numeric_source(self):
        rules = [self.rule("o.label", "sum", "total", "decimal", 1)]
        with self.assertRaises((TypeError, ValueError)):
            self.engine._aggregate(self.frame, ExtractionSpec(aggregations=rules), "summary")

    def test_template_and_repository_support_optional_aggregation_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "plan.xlsx"; create_template(path)
            workbook = load_workbook(path)
            self.assertEqual([cell.value for cell in workbook["分组字段"][1]], GROUP_HEADERS)
            self.assertEqual([cell.value for cell in workbook["聚合规则"][1]], AGGREGATION_HEADERS)
            spec = ExcelSpecRepository().load(path)
            self.assertEqual(spec.groups, []); self.assertEqual(spec.aggregations, [])
            del workbook["分组字段"]; del workbook["聚合规则"]
            legacy = Path(directory) / "legacy.xlsx"; workbook.save(legacy)
            self.assertEqual(ExcelSpecRepository().load(legacy).groups, [])

    def test_validator_rejects_invalid_aggregation_contracts(self):
        base = dict(plans=[{"任务ID": "summary", "启用": "是"}], objects=[{"任务ID": "summary", "Sheet名称": "Data", "对象别名": "o", "表头行": 1, "是否主表": "是"}])
        invalid = ExtractionSpec(**base, fields=[{"任务ID": "summary", "源字段": "o.id", "目标字段": "id", "目标类型": "integer"}],
            groups=self.groups, aggregations=[self.rule("o.amount", "count_all", "department", "bad", 0)])
        errors = SpecValidator().validate(invalid).errors
        self.assertTrue(any("不能同时配置字段映射" in error for error in errors))
        self.assertTrue(any("count_all 不应填写源字段" in error for error in errors))
        self.assertTrue(any("目标字段存在重复" in error for error in errors))
        self.assertTrue(any("目标类型" in error for error in errors))
        self.assertTrue(any("正整数" in error for error in errors))
        duplicate_groups = ExtractionSpec(**base, groups=self.groups + [{**self.groups[0], "目标字段": "department_2", "分组顺序": 2}],
                                           aggregations=[self.rule("", "count_all", "rows", "integer", 1)])
        self.assertTrue(any("分组源字段不能重复" in error for error in SpecValidator().validate(duplicate_groups).errors))

    def test_scalar_expression_is_broadcast_and_converted(self):
        frame = pd.DataFrame({"o.id": [1, 2]}, index=[4, 8])
        fields = [
            {"任务ID": "plain", "目标字段": "integer_value", "目标类型": "integer", "转换表达式": "1", "字段顺序": 1},
            {"任务ID": "plain", "目标字段": "decimal_value", "目标类型": "decimal", "转换表达式": "1.5", "字段顺序": 2},
            {"任务ID": "plain", "目标字段": "string_value", "目标类型": "string", "转换表达式": '"ok"', "字段顺序": 3},
            {"任务ID": "plain", "目标字段": "date_value", "目标类型": "datetime", "转换表达式": '"2024-01-02"', "字段顺序": 4}]
        result = self.engine._select(frame, ExtractionSpec(fields=fields), "plain")
        self.assertEqual(result["integer_value"].tolist(), [1, 1])
        self.assertEqual(result["decimal_value"].tolist(), [1.5, 1.5])
        self.assertEqual(result["string_value"].tolist(), ["ok", "ok"])
        self.assertEqual(result["date_value"].dt.strftime("%Y-%m-%d").tolist(), ["2024-01-02", "2024-01-02"])

    def test_excel_plan_to_grouped_result_end_to_end(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); plan = root / "plan.xlsx"; source = root / "source.xlsx"
            create_template(plan)
            workbook = load_workbook(plan)
            for sheet in ["抽取计划", "数据对象", "关联关系", "字段映射", "过滤条件", "分组字段", "聚合规则"]:
                ws = workbook[sheet]
                if ws.max_row > 1: ws.delete_rows(2, ws.max_row - 1)
            workbook["抽取计划"].append(["summary", "是", ""])
            workbook["数据对象"].append(["summary", "订单", "o", 1, "是", ""])
            workbook["数据对象"].append(["summary", "明细", "i", 1, "否", ""])
            workbook["关联关系"].append(["summary", 1, "LEFT JOIN", "o.order_id", "i", "i.order_id", ""])
            workbook["过滤条件"].append(["summary", 1, 1, "o.status", "=", "paid", "", ""])
            workbook["分组字段"].append(["summary", "o.customer", "customer", "string", 1, ""])
            workbook["聚合规则"].append(["summary", "", "count_all", "line_count", "integer", "", 1, ""])
            workbook["聚合规则"].append(["summary", "o.order_id", "count_distinct", "order_count", "integer", "", 2, ""])
            workbook["聚合规则"].append(["summary", "i.amount", "sum", "total", "decimal", "", 3, ""])
            workbook.save(plan)
            with pd.ExcelWriter(source) as writer:
                pd.DataFrame({"order_id": [1, 2, 3], "customer": ["A", "A", "B"], "status": ["paid", "paid", "cancelled"]}).to_excel(writer, sheet_name="订单", index=False)
                pd.DataFrame({"order_id": [1, 1, 2, 3], "amount": [10, 20, 5, 99]}).to_excel(writer, sheet_name="明细", index=False)
            spec = ExcelSpecRepository().load(plan)
            self.assertEqual(SpecValidator().validate(spec).errors, [])
            preview = ExtractionService().preview(plan, "summary")
            self.assertIn("分组字段: 1 个", preview); self.assertIn("聚合规则: 3 条", preview)
            result = self.engine.execute(spec, "summary", source)
            self.assertEqual(result.to_dict("records"), [{"customer": "A", "line_count": 3, "order_count": 2, "total": 35.0}])
            output = root / "summary.csv"
            completed = subprocess.run([sys.executable, "-m", "excelflow", "run", "--plan", str(plan), "--task", "summary",
                                        "--source", str(source), "--format", "csv", "--output", str(output)],
                                       text=True, capture_output=True, check=False)
            self.assertEqual(completed.returncode, 0, completed.stderr)
            self.assertEqual(pd.read_csv(output).to_dict("records"), [{"customer": "A", "line_count": 3, "order_count": 2, "total": 35.0}])


    def test_repository_rejects_plan_missing_required_sheets(self):
        # 计划文件缺少必需 Sheet（被改名或手改）时，必须列出缺失项，而不是抛无上下文的 KeyError
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "bad_plan.xlsx"
            wb = Workbook(); ws = wb.active; assert ws is not None
            ws.title = "抽取计划"; wb.save(path)
            with self.assertRaisesRegex(ValueError, "缺少工作表"):
                ExcelSpecRepository().load(path)


if __name__ == "__main__": unittest.main()
