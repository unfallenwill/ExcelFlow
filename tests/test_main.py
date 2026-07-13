import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excelflow import create_template, run_task, validate


class ExtractionTest(unittest.TestCase):
    def _plan(self, root: Path, output_format: str) -> Path:
        plan_path = root / "plan.xlsx"
        create_template(plan_path)
        wb = load_workbook(plan_path)
        plan = wb["抽取计划"]
        plan.delete_rows(2, plan.max_row)
        plan.append(["orders", "是", "测试任务"])
        objects = wb["数据对象"]
        objects.delete_rows(2, objects.max_row)
        objects.append(["orders", "订单", "o", 1, "是", ""])
        objects.append(["orders", "客户", "c", 1, "否", ""])
        joins = wb["关联关系"]
        joins.delete_rows(2, joins.max_row)
        joins.append(["orders", 1, "LEFT JOIN", "o.customer_id", "c", "c.customer_id", ""])
        fields = wb["字段映射"]
        fields.delete_rows(2, fields.max_row)
        fields.append(["orders", "o.id", "order_id", "integer", "", 1, ""])
        fields.append(["orders", "c.name", "customer_name", "string", "", 2, ""])
        fields.append(["orders", "o.amount", "amount", "decimal", "", 3, ""])
        filters = wb["过滤条件"]
        filters.delete_rows(2, filters.max_row)
        filters.append(["orders", 1, 1, "o.amount", ">=", 20, "", ""])
        wb.save(plan_path)
        return plan_path

    def test_excel_to_jsonl(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "订单"
            ws.append(["id", "customer_id", "amount", "updated_at"])
            ws.append([1, 10, 20, "2026-01-03"])
            customers = wb.create_sheet("客户")
            customers.append(["customer_id", "name"])
            customers.append([10, "张三"])
            wb.save(source)
            plan = self._plan(root, "jsonl")
            count, output = run_task(plan, "orders", source, "jsonl", root / "result.jsonl")
            self.assertEqual(count, 1)
            self.assertEqual(
                json.loads(output.read_text()),
                {"order_id": 1, "customer_name": "张三", "amount": 20},
            )

    def test_four_sheet_multilevel_join(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            orders = wb.active
            orders.title = "订单"
            orders.append(["id", "customer_id", "tenant", "updated_at"])
            orders.append([1, 10, "A", "2026-01-03"])
            orders.append([2, 20, "A", "2026-01-04"])
            orders.append([3, 99, "A", "2026-01-05"])
            customers = wb.create_sheet("客户")
            customers.append(["customer_id", "name", "region_id"])
            customers.append([10, "张三", 100])
            customers.append([20, "李四", 200])
            regions = wb.create_sheet("区域")
            regions.append(["region_id", "region_name"])
            regions.append([100, "华东"])
            regions.append([200, "华西"])
            items = wb.create_sheet("明细")
            items.append(["order_id", "tenant", "price"])
            items.append([1, "A", 50])
            items.append([2, "B", 100])
            wb.save(source)

            plan_path = root / "plan.xlsx"
            create_template(plan_path)
            plan_wb = load_workbook(plan_path)
            plan = plan_wb["抽取计划"]
            plan.delete_rows(2, plan.max_row)
            plan.append(["report", "是", "四表关联测试"])
            objects = plan_wb["数据对象"]
            objects.delete_rows(2, objects.max_row)
            objects.append(["report", "订单", "o", 1, "是", ""])
            objects.append(["report", "客户", "c", 1, "否", ""])
            objects.append(["report", "区域", "r", 1, "否", ""])
            objects.append(["report", "明细", "i", 1, "否", ""])
            joins = plan_wb["关联关系"]
            joins.delete_rows(2, joins.max_row)
            joins.append(["report", 1, "LEFT JOIN", "o.customer_id", "c", "c.customer_id", ""])
            joins.append(["report", 2, "INNER JOIN", "c.region_id", "r", "r.region_id", ""])
            joins.append(["report", 3, "LEFT JOIN", "o.id", "i", "i.order_id", "复合条件1"])
            joins.append(["report", 3, "LEFT JOIN", "o.tenant", "i", "i.tenant", "复合条件2"])
            fields = plan_wb["字段映射"]
            fields.delete_rows(2, fields.max_row)
            fields.append(["report", "o.id", "order_id", "integer", "", 1, ""])
            fields.append(["report", "c.name", "customer", "string", "", 2, ""])
            fields.append(["report", "r.region_name", "region", "string", "", 3, ""])
            fields.append(["report", "i.price", "price", "decimal", "", 4, ""])
            fields.append(
                ["report", "", "double_price", "decimal", "coalesce(i.price, 0) * 2", 5, ""]
            )
            filters = plan_wb["过滤条件"]
            filters.delete_rows(2, filters.max_row)
            filters.append(["report", 1, 1, "r.region_name", "IN", "华东,华西", "", ""])
            plan_wb.save(plan_path)

            self.assertTrue(validate(plan_path).ok)
            count, output = run_task(plan_path, "report", source, "jsonl", root / "result.jsonl")
            rows = [json.loads(line) for line in output.read_text().splitlines()]
            self.assertEqual(count, 2)
            self.assertEqual(
                rows,
                [
                    {
                        "order_id": 1,
                        "customer": "张三",
                        "region": "华东",
                        "price": 50,
                        "double_price": 100,
                    },
                    {
                        "order_id": 2,
                        "customer": "李四",
                        "region": "华西",
                        "price": None,
                        "double_price": 0,
                    },
                ],
            )


if __name__ == "__main__":
    unittest.main()
