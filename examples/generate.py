"""Generate the progressive ExcelFlow tutorial workbooks."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from excelflow import create_template

ROOT = Path(__file__).resolve().parent / "tutorial"


def source_workbook() -> None:
    workbook = Workbook()
    orders = workbook.active
    orders.title = "订单"
    orders.append(["order_id", "customer_id", "tenant", "status", "amount"])
    orders.append([1001, 10, "A", "paid", 727])
    orders.append([1002, 20, "A", "pending", 499])
    orders.append([1003, 20, "B", "paid", 80])
    customers = workbook.create_sheet("客户")
    customers.append(["customer_id", "customer_name"])
    customers.append([10, "张三"])
    customers.append([20, "李四"])
    items = workbook.create_sheet("订单明细")
    items.append(["order_id", "tenant", "product", "quantity", "unit_price"])
    items.append([1001, "A", "键盘", 2, 299])
    items.append([1001, "A", "鼠标", 1, 129])
    items.append([1002, "A", "显示器", 1, 499])
    items.append([1003, "A", "错误租户数据", 1, 80])
    workbook.save(ROOT / "source.xlsx")


def empty_plan(filename: str, task_id: str, note: str):
    path = ROOT / filename
    create_template(path)
    workbook = load_workbook(path)
    plan = workbook["抽取计划"]
    plan.delete_rows(2, plan.max_row)
    plan.append([task_id, "是", note])
    for name in ("数据对象", "关联关系", "字段映射", "过滤条件", "分组字段", "聚合规则"):
        sheet = workbook[name]
        sheet.delete_rows(2, sheet.max_row)
    return path, workbook


def common_order_fields(workbook, task_id: str) -> None:
    fields = workbook["字段映射"]
    fields.append([task_id, "o.order_id", "order_id", "integer", "", 1, ""])
    fields.append([task_id, "o.status", "status", "string", "", 2, ""])
    fields.append([task_id, "o.amount", "amount", "decimal", "", 3, ""])


def lesson_01() -> None:
    path, workbook = empty_plan("01_single_sheet.xlsx", "lesson_01", "单 Sheet 和字段映射")
    workbook["数据对象"].append(["lesson_01", "订单", "o", 1, "是", "主表"])
    common_order_fields(workbook, "lesson_01")
    workbook.save(path)


def lesson_02() -> None:
    path, workbook = empty_plan("02_filters.xlsx", "lesson_02", "同组 AND 过滤")
    workbook["数据对象"].append(["lesson_02", "订单", "o", 1, "是", "主表"])
    common_order_fields(workbook, "lesson_02")
    filters = workbook["过滤条件"]
    filters.append(["lesson_02", 1, 1, "o.status", "=", "paid", "", "同组条件使用 AND"])
    filters.append(["lesson_02", 1, 2, "o.amount", ">=", 100, "", ""])
    workbook.save(path)


def add_customer_join(workbook, task_id: str) -> None:
    objects = workbook["数据对象"]
    objects.append([task_id, "订单", "o", 1, "是", "主表"])
    objects.append([task_id, "客户", "c", 1, "否", "客户信息"])
    workbook["关联关系"].append(
        [task_id, 1, "LEFT JOIN", "o.customer_id", "c", "c.customer_id", ""]
    )


def lesson_03() -> None:
    path, workbook = empty_plan("03_join.xlsx", "lesson_03", "两个 Sheet 关联")
    add_customer_join(workbook, "lesson_03")
    fields = workbook["字段映射"]
    fields.append(["lesson_03", "o.order_id", "order_id", "integer", "", 1, ""])
    fields.append(["lesson_03", "c.customer_name", "customer", "string", "", 2, ""])
    fields.append(["lesson_03", "o.amount", "amount", "decimal", "", 3, ""])
    workbook.save(path)


def lesson_04() -> None:
    path, workbook = empty_plan("04_derived_columns.xlsx", "lesson_04", "多 Sheet、复合键与衍生列")
    add_customer_join(workbook, "lesson_04")
    workbook["数据对象"].append(["lesson_04", "订单明细", "i", 1, "否", "订单行"])
    joins = workbook["关联关系"]
    joins.append(["lesson_04", 2, "LEFT JOIN", "o.order_id", "i", "i.order_id", "复合键1"])
    joins.append(["lesson_04", 2, "LEFT JOIN", "o.tenant", "i", "i.tenant", "复合键2"])
    fields = workbook["字段映射"]
    for row in [
        ["lesson_04", "o.order_id", "order_id", "integer", "", 1, ""],
        ["lesson_04", "c.customer_name", "customer", "string", "", 2, ""],
        ["lesson_04", "i.product", "product", "string", "", 3, ""],
        ["lesson_04", "i.quantity", "quantity", "integer", "", 4, ""],
        ["lesson_04", "i.unit_price", "unit_price", "decimal", "", 5, ""],
        [
            "lesson_04",
            "",
            "line_amount",
            "decimal",
            "coalesce(i.quantity, 0) * coalesce(i.unit_price, 0)",
            6,
            "衍生列",
        ],
    ]:
        fields.append(row)
    filters = workbook["过滤条件"]
    filters.append(["lesson_04", 1, 1, "o.status", "=", "paid", "", ""])
    filters.append(["lesson_04", 1, 2, "o.amount", ">=", 100, "", ""])
    workbook.save(path)


def lesson_05() -> None:
    path, workbook = empty_plan("05_aggregation.xlsx", "lesson_05", "按客户汇总订单")
    add_customer_join(workbook, "lesson_05")
    workbook["分组字段"].append(
        ["lesson_05", "c.customer_name", "customer", "string", 1, "按客户分组"]
    )
    rules = workbook["聚合规则"]
    rules.append(["lesson_05", "", "count_all", "order_count", "integer", "", 1, "订单行数"])
    rules.append(["lesson_05", "o.amount", "sum", "total_amount", "decimal", "", 2, "订单总金额"])
    rules.append(
        ["lesson_05", "o.status", "concat_agg", "statuses", "string", "|", 3, "保持原顺序"]
    )
    workbook.save(path)


if __name__ == "__main__":
    ROOT.mkdir(parents=True, exist_ok=True)
    source_workbook()
    lesson_01()
    lesson_02()
    lesson_03()
    lesson_04()
    lesson_05()
    print(f"Tutorial generated in {ROOT}")
