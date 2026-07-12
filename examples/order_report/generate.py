"""Generate the source workbook and ExcelFlow plan used by this example."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from excelflow import create_template


ROOT = Path(__file__).resolve().parent


def create_source() -> None:
    workbook = Workbook()
    orders = workbook.active
    orders.title = "订单"
    orders.append(["order_id", "customer_id", "tenant", "status", "amount"])
    orders.append([1001, 10, "A", "paid", 727])
    orders.append([1002, 20, "A", "pending", 499])
    orders.append([1003, 20, "B", "paid", 80])

    customers = workbook.create_sheet("客户")
    customers.append(["customer_id", "customer_name"])
    customers.append([10, "张三"])
    customers.append([20, "李四"])

    items = workbook.create_sheet("订单明细")
    items.append(["order_id", "tenant", "product", "quantity", "unit_price"])
    items.append([1001, "A", "键盘", 2, 299])
    items.append([1001, "A", "鼠标", 1, 129])
    items.append([1002, "A", "显示器", 1, 499])
    items.append([1003, "A", "错误租户数据", 1, 80])
    workbook.save(ROOT / "source.xlsx")


def create_plan() -> None:
    path = ROOT / "plan.xlsx"
    create_template(path)
    workbook = load_workbook(path)

    plan = workbook["抽取计划"]
    plan.delete_rows(2, plan.max_row)
    plan.append(["order_report", "是", "多 Sheet 订单明细报表"])

    objects = workbook["数据对象"]
    objects.delete_rows(2, objects.max_row)
    objects.append(["order_report", "订单", "o", 1, "是", "主表"])
    objects.append(["order_report", "客户", "c", 1, "否", "客户信息"])
    objects.append(["order_report", "订单明细", "i", 1, "否", "订单行"])

    joins = workbook["关联关系"]
    joins.delete_rows(2, joins.max_row)
    joins.append(["order_report", 1, "LEFT JOIN", "o.customer_id", "c", "c.customer_id", ""])
    joins.append(["order_report", 2, "LEFT JOIN", "o.order_id", "i", "i.order_id", "复合键1"])
    joins.append(["order_report", 2, "LEFT JOIN", "o.tenant", "i", "i.tenant", "复合键2"])

    fields = workbook["字段映射"]
    fields.delete_rows(2, fields.max_row)
    fields.append(["order_report", "o.order_id", "order_id", "integer", "", 1, ""])
    fields.append(["order_report", "c.customer_name", "customer", "string", "", 2, ""])
    fields.append(["order_report", "i.product", "product", "string", "", 3, ""])
    fields.append(["order_report", "i.quantity", "quantity", "integer", "", 4, ""])
    fields.append(["order_report", "i.unit_price", "unit_price", "decimal", "", 5, ""])
    fields.append(["order_report", "", "line_amount", "decimal", "coalesce(i.quantity, 0) * coalesce(i.unit_price, 0)", 6, "衍生列"])

    filters = workbook["过滤条件"]
    filters.delete_rows(2, filters.max_row)
    filters.append(["order_report", 1, 1, "o.status", "=", "paid", "", "同组内 AND"])
    filters.append(["order_report", 1, 2, "o.amount", ">=", 100, "", ""])
    workbook.save(path)


if __name__ == "__main__":
    create_source()
    create_plan()
    print(f"Example generated in {ROOT}")
