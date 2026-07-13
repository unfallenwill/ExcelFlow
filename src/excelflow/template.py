from collections.abc import Sequence
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .schema import (
    AGGREGATION_HEADERS,
    FIELD_HEADERS,
    FILTER_HEADERS,
    GROUP_HEADERS,
    JOIN_HEADERS,
    OBJECT_HEADERS,
    PLAN_HEADERS,
)


def _style(ws: Worksheet, widths: Sequence[int]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = (
            fill,
            Font(color="FFFFFF", bold=True),
            Alignment(horizontal="center"),
        )
    ws.freeze_panes, ws.auto_filter.ref = "A2", ws.dimensions
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width


def _validation(ws: Worksheet, column: str, values: Sequence[str], limit: int = 1000) -> None:
    rule = DataValidation(type="list", formula1='"' + ",".join(values) + '"')
    ws.add_data_validation(rule)
    rule.add(f"{column}2:{column}{limit}")


def create_template(path: Path) -> None:
    wb = Workbook()
    plan = cast(Worksheet, wb.active)
    plan.title = "抽取计划"
    plan.append(PLAN_HEADERS)
    plan.append(["demo_orders", "否", "示例任务"])
    _style(plan, [18, 9, 40])
    _validation(plan, "B", ["是", "否"])
    objects = wb.create_sheet("数据对象")
    objects.append(OBJECT_HEADERS)
    objects.append(["demo_orders", "订单", "o", 1, "是", "主表"])
    objects.append(["demo_orders", "订单明细", "i", 1, "否", "关联表"])
    _style(objects, [18, 24, 16, 12, 14, 28])
    _validation(objects, "E", ["是", "否"])
    joins = wb.create_sheet("关联关系")
    joins.append(JOIN_HEADERS)
    joins.append(
        [
            "demo_orders",
            1,
            "LEFT JOIN",
            "o.order_id",
            "i",
            "i.order_id",
            "相同顺序可配置多个关联字段",
        ]
    )
    _style(joins, [18, 12, 16, 22, 16, 22, 36])
    _validation(joins, "C", ["INNER JOIN", "LEFT JOIN"])
    fields = wb.create_sheet("字段映射")
    fields.append(FIELD_HEADERS)
    fields.append(["demo_orders", "o.order_id", "order_id", "integer", "", 1, ""])
    fields.append(
        [
            "demo_orders",
            "",
            "total_amount",
            "decimal",
            "coalesce(i.quantity, 0) * coalesce(i.unit_price, 0)",
            2,
            "Pandas安全表达式",
        ]
    )
    _style(fields, [18, 22, 20, 16, 52, 12, 28])
    _validation(fields, "D", ["integer", "decimal", "string", "datetime"], 5000)
    filters = wb.create_sheet("过滤条件")
    filters.append(FILTER_HEADERS)
    filters.append(["demo_orders", 1, 1, "o.status", "=", "paid", "", "组内AND，组间OR"])
    _style(filters, [18, 12, 12, 20, 14, 24, 24, 32])
    _validation(
        filters,
        "E",
        [
            "=",
            "!=",
            ">",
            ">=",
            "<",
            "<=",
            "IN",
            "NOT IN",
            "BETWEEN",
            "LIKE",
            "NOT LIKE",
            "IS NULL",
            "IS NOT NULL",
        ],
        5000,
    )
    groups = wb.create_sheet("分组字段")
    groups.append(GROUP_HEADERS)
    _style(groups, [18, 22, 20, 16, 12, 32])
    _validation(groups, "D", ["integer", "decimal", "string", "datetime"], 5000)
    aggregations = wb.create_sheet("聚合规则")
    aggregations.append(AGGREGATION_HEADERS)
    _style(aggregations, [18, 22, 18, 20, 16, 14, 12, 32])
    _validation(
        aggregations,
        "C",
        [
            "count",
            "count_all",
            "count_distinct",
            "sum",
            "avg",
            "min",
            "max",
            "first",
            "last",
            "concat_agg",
        ],
        5000,
    )
    _validation(aggregations, "E", ["integer", "decimal", "string", "datetime"], 5000)
    guide = wb.create_sheet("填写说明")
    guide.append(["项目", "说明"])
    for row in [
        (
            "执行引擎",
            "Pandas；计划会被解释为 read_excel、merge、布尔过滤和 Series 运算，不执行 SQL 或 eval",
        ),
        ("数据对象", "每个任务必须且只能有一个主表；同一源 Excel 可配置多个 Sheet"),
        ("关联关系", "支持 INNER JOIN 和 LEFT JOIN；相同关联顺序的多行组成复合关联键"),
        ("字段", "映射、过滤和表达式字段使用 别名.字段"),
        ("转换表达式", "支持安全的数值、字符串、日期、条件函数；详见表达式参考"),
        (
            "分组聚合",
            "分组字段定义结果维度，聚合规则支持计数、求和、平均值、极值及文本合并；聚合任务不再执行字段映射",
        ),
        ("目标类型", "从下拉框选择 integer、decimal、string 或 datetime，执行时会转换输出列类型"),
        ("过滤条件", "同组内 AND，不同组之间 OR；IN 值用英文逗号分隔"),
        ("输出", "执行 run 时分别指定输出格式（csv/jsonl/xlsx）和输出路径"),
    ]:
        guide.append(row)
    _style(guide, [20, 100])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
