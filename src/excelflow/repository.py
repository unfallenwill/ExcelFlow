from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .schema import ExtractionSpec


class ExcelSpecRepository:
    """Repository adapter for the declarative Excel plan."""

    required_sheets = {"抽取计划", "数据对象", "关联关系", "字段映射", "过滤条件"}

    @staticmethod
    def _records(ws: Worksheet) -> list[dict[str, Any]]:
        headers = [str(x.value).strip() if x.value is not None else "" for x in ws[1]]
        return [
            dict(zip(headers, row, strict=True))
            for row in ws.iter_rows(min_row=2, values_only=True)
            if any(value is not None and str(value).strip() for value in row)
        ]

    def load(self, path: Path) -> ExtractionSpec:
        # read_only 工作簿持有文件句柄，必须显式关闭，否则 Windows 上临时文件清理会失败。
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            missing = self.required_sheets - set(wb.sheetnames)
            if missing:
                raise ValueError(f"缺少工作表: {', '.join(sorted(missing))}")
            return ExtractionSpec(
                plans=self._records(wb["抽取计划"]),
                objects=self._records(wb["数据对象"]),
                joins=self._records(wb["关联关系"]),
                fields=self._records(wb["字段映射"]),
                filters=self._records(wb["过滤条件"]),
                groups=self._records(wb["分组字段"]) if "分组字段" in wb.sheetnames else [],
                aggregations=self._records(wb["聚合规则"]) if "聚合规则" in wb.sheetnames else [],
            )
        finally:
            wb.close()
