from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PLAN_HEADERS = [
    "任务ID", "启用", "抽取模式", "增量字段", "开始值", "结束值",
    "输出格式", "输出路径", "批次大小", "负责人", "备注",
]
OBJECT_HEADERS = ["任务ID", "Sheet名称", "对象别名", "表头行", "是否主表", "备注"]
JOIN_HEADERS = ["任务ID", "关联顺序", "关联类型", "左侧字段", "右侧对象", "右侧字段", "备注"]
FIELD_HEADERS = ["任务ID", "源字段", "目标字段", "目标类型", "转换表达式", "字段顺序", "备注"]
FILTER_HEADERS = ["任务ID", "条件组", "条件序号", "字段", "运算符", "值1", "值2", "备注"]

SAFE_FIELD = re.compile(r"^[A-Za-z_][A-Za-z0-9_$]*$")
QUALIFIED_FIELD = re.compile(r"^(?:[A-Za-z_][A-Za-z0-9_$]*\.)?[A-Za-z_][A-Za-z0-9_$]*$")


@dataclass
class ValidationResult:
    errors: list[str]
    warnings: list[str]

    @property
    def ok(self) -> bool:
        return not self.errors


def _style_sheet(ws, widths: list[int]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width


def create_template(path: Path) -> None:
    wb = Workbook()
    plan = wb.active
    plan.title = "抽取计划"
    plan.append(PLAN_HEADERS)
    plan.append(["demo_orders", "否", "增量", "o.updated_at", "${START_TIME}", "${END_TIME}", "csv", "./output/orders.csv", 10000, "数据组", "示例任务，启用前请修改"])
    _style_sheet(plan, [18, 9, 12, 20, 20, 20, 12, 30, 12, 14, 28])

    objects = wb.create_sheet("数据对象")
    objects.append(OBJECT_HEADERS)
    objects.append(["demo_orders", "订单", "o", 1, "是", "主表"])
    objects.append(["demo_orders", "订单明细", "i", 1, "否", "关联表示例"])
    _style_sheet(objects, [18, 24, 16, 12, 14, 28])
    primary = DataValidation(type="list", formula1='"是,否"')
    objects.add_data_validation(primary); primary.add("E2:E1000")

    joins = wb.create_sheet("关联关系")
    joins.append(JOIN_HEADERS)
    joins.append(["demo_orders", 1, "LEFT JOIN", "o.order_id", "i", "i.order_id", "同一关联顺序可配置多个 ON 条件"])
    _style_sheet(joins, [18, 12, 16, 22, 16, 22, 36])
    join_type = DataValidation(type="list", formula1='"INNER JOIN,LEFT JOIN"')
    joins.add_data_validation(join_type); join_type.add("C2:C1000")

    fields = wb.create_sheet("字段映射")
    fields.append(FIELD_HEADERS)
    fields.append(["demo_orders", "o.order_id", "order_id", "integer", "", 1, ""])
    fields.append(["demo_orders", "i.amount", "amount", "decimal", "", 2, ""])
    fields.append(["demo_orders", "o.updated_at", "updated_at", "datetime", "", 3, ""])
    _style_sheet(fields, [18, 20, 20, 16, 30, 12, 28])

    filters = wb.create_sheet("过滤条件")
    filters.append(FILTER_HEADERS)
    filters.append(["demo_orders", 1, 1, "o.status", "=", "paid", "", "同组内为 AND"])
    filters.append(["demo_orders", 1, 2, "i.amount", ">=", 100, "", ""])
    filters.append(["demo_orders", 2, 1, "o.region", "IN", "华东,华南", "", "不同组之间为 OR"])
    _style_sheet(filters, [18, 12, 12, 20, 14, 24, 24, 32])
    operator = DataValidation(type="list", formula1='"=,!=,>,>=,<,<=,IN,NOT IN,BETWEEN,LIKE,NOT LIKE,IS NULL,IS NOT NULL"')
    filters.add_data_validation(operator); operator.add("E2:E5000")

    guide = wb.create_sheet("填写说明")
    guide.append(["项目", "说明"])
    rows = [
        ("使用流程", "填写任务 → validate 校验 → preview 预览 SQL → run 时指定源 Excel 执行抽取"),
        ("启用", "是/否；只有“是”的任务允许执行"),
        ("数据对象", "每个任务配置一个或多个 Sheet 和对象别名，并且必须且只能有一个主表"),
        ("关联关系", "支持 INNER JOIN 和 LEFT JOIN；相同关联顺序的多行表示多个 AND 连接条件"),
        ("抽取模式", "全量或增量；增量任务必须填写增量字段"),
        ("开始值/结束值", "增量区间为 [开始值, 结束值)；支持 ${ENV_NAME} 环境变量占位符"),
        ("过滤条件", "可选的 SQL 条件，不要填写 WHERE；内容会原样拼接，执行前务必 preview"),
        ("字段映射", "源字段使用 别名.字段；不配置时抽取 *；转换表达式是内部 SQL 表达式"),
        ("过滤条件", "一行一个条件；同一条件组内使用 AND，不同条件组之间使用 OR；没有记录表示不过滤"),
        ("条件值", "IN/NOT IN 的值1使用英文逗号分隔；BETWEEN 同时填写值1和值2；IS NULL 类无需填写值"),
        ("Excel数据源", "执行 run 命令时传入同一个工作簿；每个数据对象的表头行从 1 开始计数"),
        ("输出格式", "csv、jsonl 或 xlsx；相对路径以当前工作目录为基准"),
    ]
    for row in rows:
        guide.append(row)
    _style_sheet(guide, [20, 90])
    guide.column_dimensions["B"].width = 100

    for ws in (plan,):
        enabled = DataValidation(type="list", formula1='"是,否"')
        mode = DataValidation(type="list", formula1='"全量,增量"')
        fmt = DataValidation(type="list", formula1='"csv,jsonl,xlsx"')
        ws.add_data_validation(enabled); enabled.add("B2:B1000")
        ws.add_data_validation(mode); mode.add("C2:C1000")
        ws.add_data_validation(fmt); fmt.add("G2:G1000")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _records(ws) -> list[dict[str, Any]]:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        result.append({headers[i]: value for i, value in enumerate(row)})
    return result


def read_spec(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    required = {"抽取计划", "数据对象", "关联关系", "字段映射", "过滤条件"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(f"缺少工作表: {', '.join(sorted(missing))}")
    return (_records(wb["抽取计划"]), _records(wb["数据对象"]), _records(wb["关联关系"]),
            _records(wb["字段映射"]), _records(wb["过滤条件"]))


def validate(path: Path) -> ValidationResult:
    errors: list[str] = []
    warnings: list[str] = []
    try:
        plans, objects, joins, fields, filters = read_spec(path)
    except Exception as exc:
        return ValidationResult([str(exc)], [])

    task_ids: list[str] = []
    for row_no, plan in enumerate(plans, 2):
        task_id = str(plan.get("任务ID") or "").strip()
        task_ids.append(task_id)
        prefix = f"抽取计划第{row_no}行"
        if not task_id:
            errors.append(f"{prefix}: 任务ID不能为空")
        mode = str(plan.get("抽取模式") or "").strip()
        if mode not in {"全量", "增量"}:
            errors.append(f"{prefix}: 抽取模式必须为全量或增量")
        if mode == "增量":
            inc = str(plan.get("增量字段") or "").strip()
            if not QUALIFIED_FIELD.fullmatch(inc):
                errors.append(f"{prefix}: 增量字段不能为空且格式必须为 别名.字段")
            if plan.get("开始值") in (None, "") or plan.get("结束值") in (None, ""):
                errors.append(f"{prefix}: 增量抽取必须同时填写开始值和结束值")
        if str(plan.get("输出格式") or "").lower() not in {"csv", "jsonl", "xlsx"}:
            errors.append(f"{prefix}: 输出格式必须为 csv、jsonl 或 xlsx")
        if not str(plan.get("输出路径") or "").strip():
            errors.append(f"{prefix}: 输出路径不能为空")
        if str(plan.get("启用") or "").strip() not in {"是", "否"}:
            errors.append(f"{prefix}: 启用必须为是或否")
    if len(task_ids) != len(set(task_ids)):
        errors.append("任务ID存在重复")
    aliases_by_task: dict[str, set[str]] = {}
    for row_no, obj in enumerate(objects, 2):
        prefix = f"数据对象第{row_no}行"
        task_id = str(obj.get("任务ID") or "").strip()
        alias = str(obj.get("对象别名") or "").strip()
        if task_id not in task_ids:
            errors.append(f"{prefix}: 任务ID不存在")
        if not str(obj.get("Sheet名称") or "").strip():
            errors.append(f"{prefix}: Sheet名称不能为空")
        if not SAFE_FIELD.fullmatch(alias):
            errors.append(f"{prefix}: 对象别名必须是简单标识符")
        aliases_by_task.setdefault(task_id, set())
        if alias in aliases_by_task[task_id]:
            errors.append(f"{prefix}: 对象别名重复")
        aliases_by_task[task_id].add(alias)
        try:
            if int(obj.get("表头行") or 1) < 1:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"{prefix}: 表头行必须是正整数")
        if str(obj.get("是否主表") or "").strip() not in {"是", "否"}:
            errors.append(f"{prefix}: 是否主表必须为是或否")
    for task_id in task_ids:
        task_objects = [x for x in objects if str(x.get("任务ID") or "").strip() == task_id]
        if sum(str(x.get("是否主表") or "").strip() == "是" for x in task_objects) != 1:
            errors.append(f"任务 {task_id}: 必须且只能配置一个主表")
    for plan in plans:
        task_id = str(plan.get("任务ID") or "").strip()
        inc = str(plan.get("增量字段") or "").strip()
        if inc and "." in inc and inc.split(".", 1)[0] not in aliases_by_task.get(task_id, set()):
            errors.append(f"任务 {task_id}: 增量字段引用了不存在的对象别名")
    for row_no, join in enumerate(joins, 2):
        prefix = f"关联关系第{row_no}行"
        task_id = str(join.get("任务ID") or "").strip()
        aliases = aliases_by_task.get(task_id, set())
        if task_id not in task_ids:
            errors.append(f"{prefix}: 任务ID不存在")
        if str(join.get("关联类型") or "").strip().upper() not in {"INNER JOIN", "LEFT JOIN"}:
            errors.append(f"{prefix}: 关联类型必须为 INNER JOIN 或 LEFT JOIN")
        try:
            if int(join.get("关联顺序")) < 1:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"{prefix}: 关联顺序必须是正整数")
        right_alias = str(join.get("右侧对象") or "").strip()
        if right_alias not in aliases:
            errors.append(f"{prefix}: 右侧对象不存在")
        for name in ("左侧字段", "右侧字段"):
            value = str(join.get(name) or "").strip()
            if not QUALIFIED_FIELD.fullmatch(value) or "." not in value or value.split(".", 1)[0] not in aliases:
                errors.append(f"{prefix}: {name}必须使用有效的 别名.字段")
        right_field = str(join.get("右侧字段") or "")
        if "." in right_field and right_field.split(".", 1)[0] != right_alias:
            errors.append(f"{prefix}: 右侧字段必须属于右侧对象")
    for task_id in task_ids:
        task_objects = [x for x in objects if str(x.get("任务ID") or "").strip() == task_id]
        primary = next((str(x.get("对象别名") or "").strip() for x in task_objects if str(x.get("是否主表") or "").strip() == "是"), None)
        joined = {primary} if primary else set()
        task_joins = [x for x in joins if str(x.get("任务ID") or "").strip() == task_id]
        for order in sorted({int(x["关联顺序"]) for x in task_joins if str(x.get("关联顺序") or "").isdigit()}):
            rows = [x for x in task_joins if str(x.get("关联顺序")) == str(order)]
            right_aliases = {str(x.get("右侧对象") or "").strip() for x in rows}
            join_types = {str(x.get("关联类型") or "").strip().upper() for x in rows}
            if len(right_aliases) != 1 or len(join_types) != 1:
                errors.append(f"任务 {task_id} 关联顺序 {order}: 关联类型和右侧对象必须一致")
                continue
            right_alias = next(iter(right_aliases))
            left_aliases = {str(x.get("左侧字段") or "").split(".", 1)[0] for x in rows}
            if not left_aliases.issubset(joined):
                errors.append(f"任务 {task_id} 关联顺序 {order}: 左侧字段只能引用主表或此前已关联的对象")
            if right_alias in joined:
                errors.append(f"任务 {task_id} 关联顺序 {order}: 右侧对象已被关联")
            joined.add(right_alias)
        expected = aliases_by_task.get(task_id, set())
        if joined != expected:
            errors.append(f"任务 {task_id}: 存在未关联的数据对象: {', '.join(sorted(expected - joined))}")
    for row_no, field in enumerate(fields, 2):
        if str(field.get("任务ID") or "").strip() not in task_ids:
            errors.append(f"字段映射第{row_no}行: 任务ID不存在")
        src = str(field.get("源字段") or "").strip()
        expr = str(field.get("转换表达式") or "").strip()
        target = str(field.get("目标字段") or "").strip()
        if not expr and not QUALIFIED_FIELD.fullmatch(src):
            errors.append(f"字段映射第{row_no}行: 源字段必须为 别名.字段，或填写转换表达式")
        elif not expr and "." in src and src.split(".", 1)[0] not in aliases_by_task.get(str(field.get("任务ID") or "").strip(), set()):
            errors.append(f"字段映射第{row_no}行: 源字段引用了不存在的对象别名")
        if not SAFE_FIELD.fullmatch(target):
            errors.append(f"字段映射第{row_no}行: 目标字段格式不合法")
    operators = {"=", "!=", ">", ">=", "<", "<=", "IN", "NOT IN", "BETWEEN", "LIKE", "NOT LIKE", "IS NULL", "IS NOT NULL"}
    for row_no, condition in enumerate(filters, 2):
        prefix = f"过滤条件第{row_no}行"
        if str(condition.get("任务ID") or "").strip() not in task_ids:
            errors.append(f"{prefix}: 任务ID不存在")
        condition_field = str(condition.get("字段") or "").strip()
        if not QUALIFIED_FIELD.fullmatch(condition_field):
            errors.append(f"{prefix}: 字段格式必须为 别名.字段")
        elif "." in condition_field and condition_field.split(".", 1)[0] not in aliases_by_task.get(str(condition.get("任务ID") or "").strip(), set()):
            errors.append(f"{prefix}: 字段引用了不存在的对象别名")
        operator = str(condition.get("运算符") or "").strip().upper()
        if operator not in operators:
            errors.append(f"{prefix}: 不支持的运算符 {operator or '(空)'}")
        try:
            if int(condition.get("条件组")) < 1 or int(condition.get("条件序号")) < 1:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"{prefix}: 条件组和条件序号必须是正整数")
        if operator not in {"IS NULL", "IS NOT NULL"} and condition.get("值1") in (None, ""):
            errors.append(f"{prefix}: {operator} 必须填写值1")
        if operator == "BETWEEN" and condition.get("值2") in (None, ""):
            errors.append(f"{prefix}: BETWEEN 必须填写值2")
    return ValidationResult(errors, warnings)


ENV_PATTERN = re.compile(r"\$\{([A-Za-z_][A-Za-z0-9_]*)\}")


def _resolve(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    match = ENV_PATTERN.fullmatch(value.strip())
    if match:
        name = match.group(1)
        if name not in os.environ:
            raise ValueError(f"环境变量 {name} 未设置")
        return os.environ[name]
    return value


def _quote_field(value: str) -> str:
    parts = value.split(".")
    if not all(SAFE_FIELD.fullmatch(part) for part in parts) or len(parts) not in {1, 2}:
        raise ValueError(f"字段格式不合法: {value}")
    return ".".join(f'"{part}"' for part in parts)


def _build_filters(task_id: str, filters: list[dict[str, Any]]) -> tuple[str, list[Any]]:
    selected = [x for x in filters if str(x.get("任务ID") or "").strip() == task_id]
    groups: dict[int, list[dict[str, Any]]] = {}
    for item in selected:
        groups.setdefault(int(item["条件组"]), []).append(item)
    group_sql: list[str] = []
    params: list[Any] = []
    for group_id in sorted(groups):
        conditions: list[str] = []
        for item in sorted(groups[group_id], key=lambda x: int(x["条件序号"])):
            field = str(item["字段"]).strip()
            quoted_field = _quote_field(field)
            operator = str(item["运算符"]).strip().upper()
            if operator in {"IS NULL", "IS NOT NULL"}:
                conditions.append(f"{quoted_field} {operator}")
            elif operator in {"IN", "NOT IN"}:
                values = [x.strip() for x in str(item["值1"]).split(",") if x.strip()]
                if not values:
                    raise ValueError(f"{operator} 至少需要一个值")
                conditions.append(f'{quoted_field} {operator} ({", ".join("?" for _ in values)})')
                params.extend(_resolve(x) for x in values)
            elif operator == "BETWEEN":
                conditions.append(f"{quoted_field} BETWEEN ? AND ?")
                params.extend([_resolve(item["值1"]), _resolve(item["值2"])])
            else:
                conditions.append(f"{quoted_field} {operator} ?")
                params.append(_resolve(item["值1"]))
        if conditions:
            group_sql.append("(" + " AND ".join(conditions) + ")")
    return " OR ".join(group_sql), params


def _build_from(task_id: str, objects: list[dict[str, Any]], joins: list[dict[str, Any]]) -> str:
    task_objects = [x for x in objects if str(x.get("任务ID") or "").strip() == task_id]
    primary = next(x for x in task_objects if str(x.get("是否主表") or "").strip() == "是")
    primary_alias = str(primary["对象别名"]).strip()
    sql = f'"{primary_alias}" AS "{primary_alias}"'
    task_joins = [x for x in joins if str(x.get("任务ID") or "").strip() == task_id]
    groups: dict[int, list[dict[str, Any]]] = {}
    for item in task_joins:
        groups.setdefault(int(item["关联顺序"]), []).append(item)
    for order in sorted(groups):
        rows = groups[order]
        join_type = str(rows[0]["关联类型"]).strip().upper()
        right_alias = str(rows[0]["右侧对象"]).strip()
        if any(str(x["关联类型"]).strip().upper() != join_type or str(x["右侧对象"]).strip() != right_alias for x in rows):
            raise ValueError(f"关联顺序 {order} 的关联类型和右侧对象必须一致")
        conditions = [f'{_quote_field(str(x["左侧字段"]).strip())} = {_quote_field(str(x["右侧字段"]).strip())}' for x in rows]
        sql += f' {join_type} "{right_alias}" AS "{right_alias}" ON ' + " AND ".join(conditions)
    return sql


def build_query(plan: dict[str, Any], objects: list[dict[str, Any]], joins: list[dict[str, Any]], fields: list[dict[str, Any]], filters: list[dict[str, Any]]) -> tuple[str, list[Any]]:
    selected = [x for x in fields if str(x.get("任务ID") or "").strip() == str(plan["任务ID"]).strip()]
    selected.sort(key=lambda x: int(x.get("字段顺序") or 999999))
    if selected:
        columns = []
        for item in selected:
            expression = str(item.get("转换表达式") or "").strip() or _quote_field(str(item.get("源字段") or "").strip())
            columns.append(f'{expression} AS "{item["目标字段"]}"')
        select_clause = ", ".join(columns)
    else:
        select_clause = "*"
    task_id = str(plan["任务ID"]).strip()
    sql = f"SELECT {select_clause} FROM {_build_from(task_id, objects, joins)}"
    clauses: list[str] = []
    params: list[Any] = []
    if str(plan.get("抽取模式") or "").strip() == "增量":
        field = _quote_field(str(plan["增量字段"]).strip())
        clauses.append(f"{field} >= ? AND {field} < ?")
        params.extend([_resolve(plan["开始值"]), _resolve(plan["结束值"])])
    filter_sql, filter_params = _build_filters(task_id, filters)
    if filter_sql:
        clauses.append(f"({filter_sql})")
        params.extend(filter_params)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    return sql, params


def _find_task(path: Path, task_id: str):
    plans, objects, joins, fields, filters = read_spec(path)
    plan = next((x for x in plans if str(x.get("任务ID")) == task_id), None)
    if not plan:
        raise ValueError(f"任务不存在: {task_id}")
    return plan, objects, joins, fields, filters


def _write_rows(output: Path, fmt: str, columns: list[str], rows: Iterable[tuple[Any, ...]]) -> int:
    output.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    if fmt == "csv":
        with output.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file); writer.writerow(columns)
            for row in rows: writer.writerow(row); count += 1
    elif fmt == "jsonl":
        with output.open("w", encoding="utf-8") as file:
            for row in rows:
                file.write(json.dumps(dict(zip(columns, row)), ensure_ascii=False, default=str) + "\n"); count += 1
    else:
        wb = Workbook(write_only=True); ws = wb.create_sheet("data"); ws.append(columns)
        for row in rows: ws.append(list(row)); count += 1
        wb.save(output)
    return count


def _normalize_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bytes)):
        return value
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat(sep=" ")
    return json.dumps(value, ensure_ascii=False, default=str)


def _load_object(conn: sqlite3.Connection, workbook, obj: dict[str, Any]) -> None:
    """Load one Excel worksheet into an aliased in-memory table."""
    records: list[dict[str, Any]] = []
    object_name = str(obj["Sheet名称"]).strip()
    alias = str(obj["对象别名"]).strip()
    header_row = int(obj.get("表头行") or 1)
    if object_name not in workbook.sheetnames:
        raise ValueError(f"Excel 中不存在工作表: {object_name}")
    ws = workbook[object_name]
    rows = ws.iter_rows(min_row=header_row, values_only=True)
    try:
        headers = [str(x).strip() if x is not None else "" for x in next(rows)]
    except StopIteration:
        raise ValueError(f"Excel 工作表为空: {object_name}")
    if not all(SAFE_FIELD.fullmatch(x) for x in headers) or len(headers) != len(set(headers)):
        raise ValueError("Excel 表头必须是非空、唯一的简单字段名")
    records = [dict(zip(headers, row)) for row in rows if any(x is not None for x in row)]
    if not records:
        raise ValueError(f"源对象没有数据: {object_name}")
    columns: list[str] = []
    for record in records:
        for key in record:
            if not SAFE_FIELD.fullmatch(str(key)):
                raise ValueError(f"源字段名不合法: {key}")
            if key not in columns:
                columns.append(key)
    quoted = ", ".join(f'"{column}"' for column in columns)
    conn.execute(f'CREATE TABLE "{alias}" ({quoted})')
    placeholders = ", ".join("?" for _ in columns)
    conn.executemany(
        f'INSERT INTO "{alias}" ({quoted}) VALUES ({placeholders})',
        ([ _normalize_value(record.get(column)) for column in columns] for record in records),
    )


def run_task(path: Path, task_id: str, source_path: Path) -> tuple[int, Path]:
    result = validate(path)
    if not result.ok:
        raise ValueError("Excel 校验失败:\n" + "\n".join(result.errors))
    plan, objects, joins, fields, filters = _find_task(path, task_id)
    if str(plan.get("启用")) != "是":
        raise ValueError(f"任务 {task_id} 未启用")
    sql, params = build_query(plan, objects, joins, fields, filters)
    batch_size = int(plan.get("批次大小") or 10000)
    output = Path(str(plan["输出路径"]))
    with sqlite3.connect(":memory:") as conn:
        workbook = load_workbook(source_path, data_only=True, read_only=True)
        task_objects = [x for x in objects if str(x.get("任务ID") or "").strip() == task_id]
        for obj in task_objects:
            _load_object(conn, workbook, obj)
        cursor = conn.execute(sql, params)
        columns = [item[0] for item in cursor.description]
        def rows():
            while batch := cursor.fetchmany(batch_size):
                yield from batch
        count = _write_rows(output, str(plan["输出格式"]).lower(), columns, rows())
    return count, output


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel 驱动的数据抽取工具")
    sub = parser.add_subparsers(dest="command", required=True)
    p_template = sub.add_parser("template", help="生成 Excel 模板")
    p_template.add_argument("path", nargs="?", default="extraction_plan.xlsx")
    p_validate = sub.add_parser("validate", help="校验 Excel 声明")
    p_validate.add_argument("path")
    p_preview = sub.add_parser("preview", help="预览任务 SQL")
    p_preview.add_argument("path"); p_preview.add_argument("task_id")
    p_run = sub.add_parser("run", help="执行一个已启用任务")
    p_run.add_argument("path"); p_run.add_argument("task_id"); p_run.add_argument("source_excel")
    args = parser.parse_args()
    try:
        if args.command == "template":
            create_template(Path(args.path)); print(f"已生成: {args.path}")
        elif args.command == "validate":
            result = validate(Path(args.path))
            for item in result.errors: print(f"错误: {item}")
            for item in result.warnings: print(f"警告: {item}")
            if result.ok: print("校验通过")
            return 0 if result.ok else 1
        elif args.command == "preview":
            plan, objects, joins, fields, filters = _find_task(Path(args.path), args.task_id)
            sql, params = build_query(plan, objects, joins, fields, filters)
            print(sql); print("参数:", params)
        elif args.command == "run":
            count, output = run_task(Path(args.path), args.task_id, Path(args.source_excel))
            print(f"抽取完成: {count} 行 -> {output}")
    except Exception as exc:
        print(f"失败: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
